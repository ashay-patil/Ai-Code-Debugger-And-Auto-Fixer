import os
import sys
import argparse
import json
import re
import shutil
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import threading
from rich.console import Console
from rich.markdown import Markdown
import google.generativeai as genai
try:
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox
    from tkinter.scrolledtext import ScrolledText
except Exception:
    tk = None


console = Console()
MODEL_NAME = os.getenv("GEMINI_MODEL_NAME", "models/gemini-2.5-flash")
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

try:
    MODEL = genai.GenerativeModel(MODEL_NAME)
except Exception:
    MODEL = None

SUPPORTED_SUFFIXES = (".py", ".js", ".jsx", ".ts", ".java", ".cpp", ".html", ".css")


def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def print_logo():
    console.print("""
[bold cyan]
  █████╗ ███████╗      ██████╗   ██████╗  ██████╗  ███████╗    ██████╗  ███████╗██████╗ ██╗   ██╗ ██████╗  ██████╗ ██████╗ ███████
██╔══██╗   ██║       ██╔════╝  ██╔══  ██╗  ██╔══██╗ ██╔════╝   ██╔══██╗ ██╔════╝██╔══██╗██║   ██║██╔════╝ ██╔════╝ ██╔════╝██╔══██╗
███████║   ██║       ██║       ██║    ██║  ██║  ██║ █████╗     ██   ██╔ ███████ ██████╔╝██║   ██║██║  ███╗██║  ███╗█████╗  ██████╔╝
██╔══██║   ██║       ██║       ██║    ██║  ██║  ██║ ██╔══╝     ██╔══██╗ ██╔══╝  ██╔══██╗██║   ██║██║   ██║██║   ██║██╔══╝  ██╔══██╗
██║  ██║ ███████╗     ╚██████╗  ╚██████╔╝ ██████╔╝ ███████╗    ██████╔╝ ███████╗██████╔╝╚██████╔╝╚██████╔╝╚██████╔╝███████╗██║  ██║
[bold magenta]🤖 AI-Based Multi-File Code Debugger + Auto Fixer[/bold magenta]
""")

def get_code_files(directory: str) -> List[str]:
    code_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(SUPPORTED_SUFFIXES):
                path = os.path.join(root, file)
                if any(ignored in Path(path).parts for ignored in ("node_modules", ".git", "__pycache__", "dist", "build")):
                    continue
                code_files.append(path)
    code_files.sort()
    return code_files

def read_file(file_path: str) -> Optional[str]:
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        console.print(f"[red]Error reading {file_path}: {e}[/red]")
        return None

def safe_backup(file_path: str):
    try:
        bak = file_path + ".bak"
        if not os.path.exists(bak) and os.path.exists(file_path):
            shutil.copyfile(file_path, bak)
    except Exception as e:
        console.print(f"[yellow]Warning: Could not create backup for {file_path}: {e}[/yellow]")

def clean_code_fences(text: str) -> str:
    if not text:
        return ""
    s = text.strip()
    s = re.sub(r"^```[\w+-]*\n?", "", s)
    s = re.sub(r"\n?```$", "", s)
    s = re.sub(r'^\s*(Fixed code|Corrected code|Suggested fix)\s*:\s*', '', s, flags=re.IGNORECASE)
    return s.strip() + "\n"

def write_file(file_path: str, new_code: str):
    try:
        safe_backup(file_path)
        cleaned = clean_code_fences(new_code)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(cleaned)
        console.print(f"[green]✅ Updated: {file_path}[/green]")
    except Exception as e:
        console.print(f"[red]❌ Error writing {file_path}: {e}[/red]")

def build_file_section(file_path: str, code: str) -> str:
    suffix = Path(file_path).suffix.lstrip('.') or 'txt'
    return f"### FILE PATH: {file_path}\n```{suffix}\n{code}\n```\n\n"

def chunk_files(files: List[str], file_contents: Dict[str, str], max_chars: int) -> Tuple[List[str], List[List[str]]]:
    batches = []
    batch_files = []
    current, current_files, length = [], [], 0
    for f in files:
        section = build_file_section(f, file_contents[f])
        if current and (length + len(section) > max_chars):
            batches.append("\n".join(current))
            batch_files.append(current_files)
            current, current_files, length = [section], [f], len(section)
        else:
            current.append(section)
            current_files.append(f)
            length += len(section)
    if current:
        batches.append("\n".join(current))
        batch_files.append(current_files)
    return batches, batch_files

def generate_content(prompt: str) -> str:
    if MODEL is None:
        raise RuntimeError("Gemini model unavailable. Check API key or model name.")
    try:
        resp = MODEL.generate_content(prompt)
        return resp.text or ""
    except Exception as e:
        return f"Error: {e}"

def review_project(all_code: str, user_prompt: Optional[str]) -> str:
    prompt = f"""
You are an expert software debugger and code reviewer.

Analyze the entire project for defects and risks. Each section below starts with "### FILE PATH:"
and contains the full code for that file. Provide precise, actionable findings and corrected code.

Scope of checks (be thorough and conservative):
- Syntax errors and runtime errors (undefined names, indexing, attribute, import errors)
- Incorrect logic and edge cases (off-by-one, null/None checks, empty input handling)
- Exception handling quality (missing try/except where needed; overbroad excepts; swallowed errors)
- Resource handling (files, network, subprocesses) and leaks; context managers where applicable
- Dependency/import issues (wrong filenames/paths/extensions, missing modules, circular imports)
- API contract mismatches (wrong params/return types, missing awaits, async misuse)
- Type issues and nullability; mutation of shared state; concurrency/async pitfalls
- HTML/CSS/JS linkage mistakes (wrong script/style paths, case sensitivity, extension mismatches)
- Frontend errors (DOM selectors, event binding, missing assets, incorrect MIME/rel attributes)
- Security concerns (eval/exec, command injection, XSS/CSRF risks, hardcoded secrets, plaintext keys)
- Performance hotspots (N+1 queries/polls, excessive loops, unbounded recursion, heavy sync I/O on UI)
- Code quality (duplication, long functions, poor naming, missing docs/comments where non-obvious)
- During api fetching the code should be wrapped in try catch block.If not then the corrected code which you will give should contain that try catch block
Additionally, include three Markdown tables for every file:

Chart 1 - Review Checklist: for the present code and not about the corrected code. The charts should be made for the existing codes
| Review Aspect | Status |
|---|---|
| Variable naming | ✅ / ❌ |
| Hardcoded values/secrets | ✅ / ❌ |
| Code repetition | ✅ / ❌ |
| Modularity | ✅ / ❌ |
| Complexity (high/med/low) | high/med/low |
| Comments & docs | ✅ / ❌ |
| Exception handling present | ✅ / ❌ |
| Dependency/import correctness | ✅ / ❌ |
| Security concerns | ✅ / ❌ |

Chart 2 - API Calls Summary (if any found):
| API Endpoint | Request (sample) | Response (sample) |
|---|---|---|
| (list endpoints or 'None') | | |

Chart 3 - Security Issues / Best Practices:
| Category | Recommendation |
|---|---|
| (e.g. plaintext password storage) | (recommendation) |

User requirement: {user_prompt or 'N/A'}
Important: Fix file import/link mismatches robustly. Example: if the folder has script.js but HTML imports app.js,
change it to script.js; similarly styles.css vs style.css, respecting actual files present. NEVER invent new files
that do not exist—choose the most plausible correct existing path/filename.

Here is the project:
{all_code}

Output format (repeat for all files whether there exists error or not. If no error is present then say that no error and in fixed code place the original code as it is, and if there is error then tell the error and give the corrected code in fixed code):
File: <path>
Error: <short description>
Fixed code:
```<language>
<corrected file code>
```

"""
    return generate_content(prompt)

def auto_fix_project(path: str, review_output: str, files_to_process: List[str], apply_all=False, interactive: bool = True, prompt_func=None, ui_logger=None):
    model = genai.GenerativeModel(MODEL_NAME)

    for file_path in files_to_process:
        code = read_file(file_path)
        if not code:
            continue

        fix_prompt = f"""
You are an AI assistant. Using the review below, fix only issues related to this file.

Review:
{review_output}

File: {file_path}
Current Code:
```{Path(file_path).suffix[1:]}
{code}
```

Output the corrected code for this file only.
"""
        try:
            resp = model.generate_content(fix_prompt)
            new_code = resp.text.strip()
            console.rule(f"[bold yellow]🧠 Suggested Fix for {file_path}[/bold yellow]")
            if ui_logger:
                ui_logger(f"\n===== Suggested Fix for {file_path} =====\n")
                ui_logger(new_code[:3000] + ("\n...[truncated]\n" if len(new_code) > 3000 else "\n"))
            else:
                console.print(Markdown(new_code[:3000]))

            if apply_all:
                write_file(file_path, new_code)
            else:
                if interactive:
                    if ui_logger:
                        apply_change = bool(prompt_func(f"Apply suggested changes to {file_path}?"))
                        if apply_change:
                            write_file(file_path, new_code)
                        else:
                            if ui_logger:
                                ui_logger(f"Skipped: {file_path}\n")
                            else:
                                console.print(f"[yellow]⏩ Skipped: {file_path}[/yellow]")
                    else:
                        choice = input(f"Apply suggested changes to {file_path}? (y/n): ").strip().lower()
                        if choice == 'y':
                            write_file(file_path, new_code)
                        else:
                            console.print(f"[yellow]⏩ Skipped: {file_path}[/yellow]")
                else:
                    if ui_logger:
                        ui_logger(f"Skipped (non-interactive): {file_path}\n")
                    else:
                        console.print(f"[yellow]⏩ Skipped (non-interactive): {file_path}[/yellow]")
        except Exception as e:
            console.print(f"[red]Error fixing {file_path}: {e}[/red]")


def _strip_inline_md(text: str) -> str:
    """Remove bold (**), italic (*), and backtick markers from a string."""
    text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)
    text = re.sub(r'\*(.+?)\*', r'\1', text)
    text = re.sub(r'`(.+?)`', r'\1', text)
    return text.strip()


def _parse_md_tables(text: str):
    """
    Yield (headers_list, rows_list) for every markdown table found in *text*.
    Rows is a list of lists (one inner list per data row).
    """
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        # A table header line has at least two pipe characters
        if line.count('|') >= 2:
            # Check that the very next line is a separator (---|--- etc.)
            if i + 1 < len(lines) and re.match(r'^\s*\|?[\s\-:|]+\|', lines[i + 1]):
                headers = [_strip_inline_md(c) for c in line.split('|') if c.strip()]
                i += 2  # skip header + separator
                rows = []
                while i < len(lines) and lines[i].count('|') >= 2:
                    cells = [_strip_inline_md(c) for c in lines[i].split('|') if c.strip() != '']
                    # Pad / trim to header length
                    while len(cells) < len(headers):
                        cells.append('')
                    rows.append(cells[:len(headers)])
                    i += 1
                yield headers, rows
                continue
        i += 1


def _parse_file_sections(text: str):
    """
    Split review text into per-file sections.
    Each section is a dict: {file, error_lines, fixed_code, tables, body_text}
    """
    # Patterns that mark the start of a new file block
    FILE_RE = re.compile(r'^(?:###\s*FILE\s*PATH\s*:|File\s*:)\s*(.+)', re.IGNORECASE)

    sections = []
    current = None

    def new_section(path):
        return {"file": path, "error_lines": [], "fixed_code": [], "tables": [], "body_text": []}

    lines = text.splitlines()
    i = 0
    in_code = False
    capturing_fixed = False
    code_buf = []

    while i < len(lines):
        line = lines[i]
        m = FILE_RE.match(line)

        if m:
            # Close previous section
            if current:
                sections.append(current)
            current = new_section(m.group(1).strip())
            in_code = False
            capturing_fixed = False
            i += 1
            continue

        if current is None:
            i += 1
            continue

        # Code fence toggle
        if line.strip().startswith("```"):
            if not in_code:
                in_code = True
                code_buf = []
            else:
                in_code = False
                if capturing_fixed:
                    current["fixed_code"] = code_buf[:]
                    capturing_fixed = False
                code_buf = []
            i += 1
            continue

        if in_code:
            code_buf.append(line)
            i += 1
            continue

        # "Fixed code:" label
        if re.match(r'^\s*(Fixed code|Corrected code)\s*:', line, re.IGNORECASE):
            capturing_fixed = True
            i += 1
            continue

        current["body_text"].append(line)
        i += 1

    if current:
        sections.append(current)

    # Extract tables from body_text for each section
    for sec in sections:
        body = "\n".join(sec["body_text"])
        sec["tables"] = list(_parse_md_tables(body))

    return sections


def save_to_excel(review_output: str, filename: str = "project_review.xlsx"):
    """
    Save Gemini review output into a properly formatted Excel workbook.
    - Sheet per file: shows errors, review checklist table, API table, security table
    - Summary sheet: one row per file with key checklist flags
    - Raw sheet: full markdown text for reference
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        console.print("[yellow]Warning: openpyxl not installed. Cannot save to Excel.[/yellow]")
        return

    # ── colour palette ──────────────────────────────────────────────
    CLR_HEADER_BG  = "2B2D42"   # dark navy
    CLR_HEADER_FG  = "FFFFFF"
    CLR_FILE_BG    = "EEF2F7"   # light blue-grey
    CLR_TABLE_HDR  = "4472C4"   # Excel blue
    CLR_TABLE_HDR_FG = "FFFFFF"
    CLR_ALT_ROW    = "DCE6F1"
    CLR_ERROR_BG   = "FDECEA"   # soft red
    CLR_OK_BG      = "E8F5E9"   # soft green
    CLR_SECTION_BG = "FFF3CD"   # soft amber

    thin = Side(style="thin", color="BDBDBD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(size=11, bold=True, color=CLR_HEADER_FG):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def body_font(size=10, bold=False, color="000000"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def write_table(ws, start_row, headers, rows, title=None):
        """Write a formatted table into ws starting at start_row. Returns next free row."""
        row = start_row
        if title:
            ws.cell(row, 1, title).font = Font(name="Arial", size=10, bold=True, color="2B2D42")
            ws.cell(row, 1).fill = fill(CLR_SECTION_BG)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(len(headers), 1))
            row += 1

        # Header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.font = hdr_font(size=10)
            c.fill = fill(CLR_TABLE_HDR)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        row += 1

        # Data rows
        for ri, data_row in enumerate(rows):
            bg = CLR_ALT_ROW if ri % 2 == 1 else "FFFFFF"
            for ci, val in enumerate(data_row, 1):
                c = ws.cell(row, ci, val)
                c.font = body_font()
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = border
                # Colour status cells
                if ci == len(headers) and len(headers) == 2:
                    if val in ("✅", "✓", "OK", "ok"):
                        c.fill = fill("C8E6C9")
                    elif val in ("❌", "✗", "FAIL", "fail"):
                        c.fill = fill("FFCDD2")
                    else:
                        c.fill = fill(bg)
                else:
                    c.fill = fill(bg)
            row += 1

        return row + 1   # blank spacer row

    # ── parse ────────────────────────────────────────────────────────
    sections = _parse_file_sections(review_output)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    TABLE_TITLES = ["Review Checklist", "API Calls Summary", "Security / Best Practices"]

    # ── per-file sheets ──────────────────────────────────────────────
    for sec in sections:
        # Safe sheet name (max 31 chars, no special chars)
        raw_name = Path(sec["file"]).name if sec["file"] else "Unknown"
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', raw_name)[:31]
        ws = wb.create_sheet(title=safe_name)
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 40

        cur_row = 1

        # ── file path banner ─────────────────────────────────────────
        ws.row_dimensions[cur_row].height = 22
        c = ws.cell(cur_row, 1, f"📄  {sec['file']}")
        c.font = Font(name="Arial", size=11, bold=True, color=CLR_HEADER_FG)
        c.fill = fill(CLR_HEADER_BG)
        c.alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)
        cur_row += 2

        # ── errors / description ──────────────────────────────────────
        error_lines = [
            l for l in sec["body_text"]
            if l.strip() and not l.strip().startswith(('|', '#', '-', '`', 'Fixed', 'Corrected'))
        ]
        # Clean markdown from error lines
        error_clean = [_strip_inline_md(re.sub(r'^\d+\.\s*', '', l)) for l in error_lines if l.strip()]
        error_clean = [e for e in error_clean if e]

        if error_clean:
            c = ws.cell(cur_row, 1, "Issues Found")
            c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            c.fill = fill("C62828")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)
            cur_row += 1
            for line in error_clean[:20]:   # cap at 20 lines
                c = ws.cell(cur_row, 1, line)
                c.font = body_font()
                c.fill = fill(CLR_ERROR_BG)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.border = border
                ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)
                ws.row_dimensions[cur_row].height = 40
                cur_row += 1
            cur_row += 1
        else:
            c = ws.cell(cur_row, 1, "✅  No errors found")
            c.font = Font(name="Arial", size=10, bold=True, color="1B5E20")
            c.fill = fill(CLR_OK_BG)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=3)
            cur_row += 2

        # ── tables ────────────────────────────────────────────────────
        for ti, (headers, rows) in enumerate(sec["tables"][:3]):
            title = TABLE_TITLES[ti] if ti < len(TABLE_TITLES) else f"Table {ti + 1}"
            # Expand column widths for this table
            for ci in range(len(headers)):
                col_letter = get_column_letter(ci + 1)
                ws.column_dimensions[col_letter].width = max(
                    ws.column_dimensions[col_letter].width, 28
                )
            cur_row = write_table(ws, cur_row, headers, rows, title=title)

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.column_dimensions["A"].width = 40
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_sum.column_dimensions[col].width = 18

    # Title
    c = ws_sum.cell(1, 1, "🔍  Project Review Summary")
    c.font = Font(name="Arial", size=13, bold=True, color=CLR_HEADER_FG)
    c.fill = fill(CLR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 26
    total_cols = 1 + 9  # file + 9 checklist aspects
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Column headers
    summary_headers = [
        "File", "Variable Naming", "Hardcoded Secrets", "Code Repetition",
        "Modularity", "Complexity", "Comments & Docs",
        "Exception Handling", "Import Correctness", "Security"
    ]
    for ci, h in enumerate(summary_headers, 1):
        c = ws_sum.cell(2, ci, h)
        c.font = hdr_font(size=10)
        c.fill = fill(CLR_TABLE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws_sum.row_dimensions[2].height = 30

    CHECKLIST_ASPECTS = [
        "variable naming", "hardcoded", "code repetition",
        "modularity", "complexity", "comments", "exception", "import", "security"
    ]

    for ri, sec in enumerate(sections, 3):
        # Col A: file name
        c = ws_sum.cell(ri, 1, Path(sec["file"]).name if sec["file"] else sec["file"])
        c.font = body_font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = border
        c.fill = fill(CLR_FILE_BG)

        # Find the Review Checklist table (first table, 2-column)
        checklist_dict = {}
        for headers, rows in sec["tables"]:
            if len(headers) == 2 and "aspect" in headers[0].lower():
                for row in rows:
                    checklist_dict[row[0].lower()] = row[1] if len(row) > 1 else ""
                break

        for ci, aspect in enumerate(CHECKLIST_ASPECTS, 2):
            val = ""
            for key, v in checklist_dict.items():
                if aspect in key.lower():
                    val = v
                    break
            c = ws_sum.cell(ri, ci, val)
            c.font = body_font(size=11)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            if val in ("✅", "✓"):
                c.fill = fill("C8E6C9")
            elif val in ("❌", "✗"):
                c.fill = fill("FFCDD2")
            else:
                c.fill = fill("FFF9C4")

    # ── Raw Review sheet ─────────────────────────────────────────────
    ws_raw = wb.create_sheet(title="Raw Review")
    ws_raw.cell(1, 1, "Gemini Review (Raw Markdown)").font = Font(name="Arial", bold=True, size=11)
    ws_raw.column_dimensions["A"].width = 130
    for ri, line in enumerate(review_output.splitlines(), 3):
        ws_raw.cell(ri, 1, line).font = Font(name="Courier New", size=9)

    wb.save(filename)
    console.print(f"[green]📊 Review saved to Excel: {filename}[/green]")


def build_markdown_renderer(out_text_widget):
    """
    Returns a render_markdown(text) function bound to *out_text_widget*.
    Handles: headers, fenced code blocks, markdown tables, inline code,
    bold/italic text, horizontal rules, and plain paragraphs.
    """
    w = out_text_widget

    # Tag configuration
    try:
        w.tag_configure("h1",      font=("Segoe UI", 13, "bold"), foreground="#1a237e", spacing1=6, spacing3=4)
        w.tag_configure("h2",      font=("Segoe UI", 12, "bold"), foreground="#283593", spacing1=5, spacing3=3)
        w.tag_configure("h3",      font=("Segoe UI", 11, "bold"), foreground="#1d4ed8", spacing1=4, spacing3=2)
        w.tag_configure("code_block", font=("Consolas", 9),  background="#eef2f7", foreground="#1a1a1a", spacing1=2, spacing3=2)
        w.tag_configure("inline_code", font=("Consolas", 9), background="#eef2f7")
        w.tag_configure("bold",    font=("Segoe UI", 10, "bold"))
        w.tag_configure("italic",  font=("Segoe UI", 10, "italic"))
        w.tag_configure("table_hdr",  font=("Segoe UI", 9, "bold"), background="#4472C4", foreground="#ffffff")
        w.tag_configure("table_odd",  font=("Segoe UI", 9),         background="#DCE6F1")
        w.tag_configure("table_even", font=("Segoe UI", 9),         background="#ffffff")
        w.tag_configure("table_sep",  font=("Consolas", 8),         foreground="#9e9e9e")
        w.tag_configure("rule",    foreground="#9e9e9e")
        w.tag_configure("info",    foreground="#2563eb")
        w.tag_configure("success", foreground="#059669")
        w.tag_configure("warning", foreground="#d97706")
        w.tag_configure("error",   foreground="#dc2626")
        w.tag_configure("section", font=("Segoe UI", 10, "bold"), foreground="#1d4ed8")
        w.tag_configure("muted",   foreground="#475569")
    except Exception:
        pass

    def _insert_inline(text_line, default_tag=None):
        """Insert a line with inline bold/italic/code formatting."""
        # Pattern: **bold**, *italic*, `code`
        pattern = re.compile(r'(\*\*(.+?)\*\*|\*(.+?)\*|`([^`]+)`)')
        pos = 0
        for m in pattern.finditer(text_line):
            # insert plain text before this match
            before = text_line[pos:m.start()]
            if before:
                w.insert("end", before, (default_tag,) if default_tag else ())
            full = m.group(0)
            if full.startswith("**"):
                w.insert("end", m.group(2), ("bold",))
            elif full.startswith("*"):
                w.insert("end", m.group(3), ("italic",))
            elif full.startswith("`"):
                w.insert("end", m.group(4), ("inline_code",))
            pos = m.end()
        tail = text_line[pos:]
        if tail:
            w.insert("end", tail, (default_tag,) if default_tag else ())

    def render_markdown(text: str):
        lines = text.splitlines()
        i = 0
        in_code = False
        code_lang = ""
        code_lines = []

        while i < len(lines):
            line = lines[i]

            # ── fenced code block ──────────────────────────────────────
            if line.strip().startswith("```"):
                if not in_code:
                    in_code = True
                    code_lang = line.strip()[3:].strip()
                    code_lines = []
                else:
                    in_code = False
                    block = "\n".join(code_lines)
                    w.insert("end", f"[{code_lang or 'code'}]\n" if code_lang else "", ("muted",))
                    w.insert("end", block + "\n", ("code_block",))
                    w.insert("end", "\n")
                i += 1
                continue

            if in_code:
                code_lines.append(line)
                i += 1
                continue

            # ── headers ───────────────────────────────────────────────
            if line.startswith("### "):
                w.insert("end", line[4:] + "\n", ("h3",))
                i += 1
                continue
            if line.startswith("## "):
                w.insert("end", line[3:] + "\n", ("h2",))
                i += 1
                continue
            if line.startswith("# "):
                w.insert("end", line[2:] + "\n", ("h1",))
                i += 1
                continue

            # ── horizontal rule ───────────────────────────────────────
            if re.match(r'^[-*_]{3,}\s*$', line):
                w.insert("end", "─" * 70 + "\n", ("rule",))
                i += 1
                continue

            # ── markdown table ────────────────────────────────────────
            if line.count('|') >= 2 and i + 1 < len(lines) and re.match(r'^\s*\|?[\s\-:|]+\|', lines[i + 1]):
                headers = [_strip_inline_md(c) for c in line.split('|') if c.strip()]
                i += 2  # skip separator row

                # Build a fixed-width column layout
                col_width = max(16, min(30, (70 // max(len(headers), 1))))

                # Header row
                header_row = " | ".join(h[:col_width].ljust(col_width) for h in headers)
                w.insert("end", header_row + "\n", ("table_hdr",))
                sep_row = "-+-".join("-" * col_width for _ in headers)
                w.insert("end", sep_row + "\n", ("table_sep",))

                row_index = 0
                while i < len(lines) and lines[i].count('|') >= 2:
                    cells = [_strip_inline_md(c) for c in lines[i].split('|') if c.strip() != ""]
                    while len(cells) < len(headers):
                        cells.append("")
                    cells = cells[:len(headers)]
                    row_str = " | ".join(c[:col_width].ljust(col_width) for c in cells)
                    tag = "table_odd" if row_index % 2 == 0 else "table_even"
                    w.insert("end", row_str + "\n", (tag,))
                    row_index += 1
                    i += 1

                w.insert("end", "\n")
                continue

            # ── numbered / bullet list items ──────────────────────────
            if re.match(r'^\s*(\d+\.|[-*])\s+', line):
                clean = re.sub(r'^\s*(\d+\.|[-*])\s+', '  • ', line)
                _insert_inline(clean)
                w.insert("end", "\n")
                i += 1
                continue

            # ── plain text (with inline formatting) ───────────────────
            if line.strip():
                _insert_inline(line)
                w.insert("end", "\n")
            else:
                w.insert("end", "\n")
            i += 1

    return render_markdown


def run_pipeline(path: str, export_json: bool, autofix: bool, apply_all: bool, userprompt: Optional[str], max_chars: int, ui_logger=None, excel_filename: str = "project_review.xlsx", interactive: bool = False, prompt_func=None):
    if ui_logger:
        ui_logger("Starting analysis...\n")
    else:
        clear_screen()
        print_logo()

    if not os.path.exists(path):
        msg = f"Invalid path: {path}"
        if ui_logger:
            ui_logger(msg + "\n")
        else:
            console.print(f"[red]{msg}[/red]")
        return

    files = get_code_files(path)
    if not files:
        if ui_logger:
            ui_logger("No source files found!\n")
        else:
            console.print("[red]No source files found![/red]")
        return

    if ui_logger:
        ui_logger(f"Analyzing {len(files)} files...\n")
    else:
        console.rule(f"[bold magenta]🔍 Analyzing {len(files)} files...[/bold magenta]")

    contents = {f: (read_file(f) or "") for f in files}
    batches, batch_files = chunk_files(files, contents, max_chars)

    all_reviews = []
    for i, (batch, batch_file_list) in enumerate(zip(batches, batch_files), 1):
        if ui_logger:
            ui_logger(f"Processing batch {i}/{len(batches)} (Review)...\n")
        else:
            console.print(f"[green]Processing batch {i}/{len(batches)} (Review)...[/green]")
        output = review_project(batch, userprompt)
        all_reviews.append(output)

        if ui_logger:
            ui_logger(f"\n==== Review for batch {i}/{len(batches)} ====\n")
            review_content = output[:20000] + ("\n...[truncated]\n" if len(output) > 20000 else "\n")
            ui_logger(review_content, is_markdown=True)
        else:
            console.rule(f"[bold blue]🧩 Review for batch {i}/{len(batches)}[/bold blue]")
            console.print(Markdown(output[:20000]))

        if autofix:
            if ui_logger:
                ui_logger(f"Processing batch {i}/{len(batches)} (Auto-Fix)...\n")
            else:
                console.rule(f"[bold green]🔧 Auto-Fix for batch {i}/{len(batches)}[/bold green]")
            auto_fix_project(path, output, batch_file_list, apply_all=apply_all, interactive=True, prompt_func=prompt_func, ui_logger=ui_logger)

    review_output = "\n\n".join(all_reviews)
    save_to_excel(review_output, filename=excel_filename)

    if export_json:
        with open('project_review.json', 'w', encoding='utf-8') as f:
            json.dump({"review": review_output}, f, indent=2)
        if ui_logger:
            ui_logger("Review saved to project_review.json\n")
        else:
            console.print("[green]📦 Review saved to project_review.json[/green]")

    if ui_logger:
        ui_logger("\nDone.\n")
    else:
        console.rule("[bold green]✅ Done[/bold green]")
    return review_output


def launch_ui():
    if tk is None:
        console.print("[red]Tkinter is not available in this environment. Please run via CLI.[/red]")
        sys.exit(1)

    app = tk.Tk()
    app.title("AI Code Debugger + Auto Fixer (Gemini)")
    app.geometry("900x700")
    try:
        app.iconbitmap(default="")
    except Exception:
        pass
    app.configure(bg="#f5f7fb")

    path_var = tk.StringVar()
    json_var = tk.BooleanVar(value=False)
    autofix_var = tk.BooleanVar(value=False)
    apply_all_var = tk.BooleanVar(value=False)
    userprompt_var = tk.StringVar()
    maxchars_var = tk.IntVar(value=200_000)
    is_running = tk.BooleanVar(value=False)

    header = tk.Frame(app, bg="#2b2d42", height=52)
    header.pack(fill=tk.X, side=tk.TOP)
    title_lbl = tk.Label(header, text="AI Code Debugger + Auto Fixer (Gemini)", fg="white", bg="#2b2d42", font=("Segoe UI", 12, "bold"))
    title_lbl.pack(side=tk.LEFT, padx=12, pady=10)
    status_lbl = tk.Label(header, text="Idle", fg="#a8dadc", bg="#2b2d42", font=("Segoe UI", 10))
    status_lbl.pack(side=tk.RIGHT, padx=12)

    frm = ttk.Frame(app, padding=10)
    frm.pack(fill=tk.BOTH, expand=True)

    path_row = ttk.Frame(frm)
    path_row.pack(fill=tk.X, pady=5)
    ttk.Label(path_row, text="Project folder:").pack(side=tk.LEFT)
    path_entry = ttk.Entry(path_row, textvariable=path_var)
    path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    def browse():
        d = filedialog.askdirectory()
        if d:
            path_var.set(d)
    browse_btn = ttk.Button(path_row, text="Browse...", command=browse)
    browse_btn.pack(side=tk.LEFT)

    opts_row1 = ttk.Frame(frm)
    opts_row1.pack(fill=tk.X, pady=5)
    json_chk = ttk.Checkbutton(opts_row1, text="Export review as JSON", variable=json_var)
    json_chk.pack(side=tk.LEFT, padx=5)
    autofix_chk = ttk.Checkbutton(opts_row1, text="Auto-fix", variable=autofix_var)
    autofix_chk.pack(side=tk.LEFT, padx=5)
    apply_all_chk = ttk.Checkbutton(opts_row1, text="Apply all fixes automatically", variable=apply_all_var)
    apply_all_chk.pack(side=tk.LEFT, padx=5)

    prompt_row = ttk.Frame(frm)
    prompt_row.pack(fill=tk.X, pady=5)
    ttk.Label(prompt_row, text="User requirement (optional):").pack(side=tk.LEFT)
    prompt_entry = ttk.Entry(prompt_row, textvariable=userprompt_var)
    prompt_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

    max_row = ttk.Frame(frm)
    max_row.pack(fill=tk.X, pady=5)
    ttk.Label(max_row, text="Max chars per batch:").pack(side=tk.LEFT)
    max_entry = ttk.Entry(max_row, textvariable=maxchars_var, width=12)
    max_entry.pack(side=tk.LEFT, padx=5)

    out_label = ttk.Label(frm, text="Output:")
    out_label.pack(anchor="w", pady=(10, 0))
    btn_row = ttk.Frame(frm)
    btn_row.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

    out_container = ttk.Frame(frm)
    out_container.pack(fill=tk.BOTH, expand=True)
    out_text = ScrolledText(
        out_container,
        height=25,
        wrap=tk.WORD,
        bg="#f8fafc",
        fg="#0f172a",
        insertbackground="#0f172a",
        borderwidth=1,
        relief="solid"
    )
    try:
        out_text.configure(font=("Consolas", 10), padx=14, pady=12, spacing1=4, spacing3=4)
    except Exception:
        pass
    out_text.pack(fill=tk.BOTH, expand=True)

    # Build the markdown renderer bound to out_text
    render_markdown = build_markdown_renderer(out_text)

    progress_placeholder = ttk.Label(btn_row, width=28)
    progress_placeholder.pack(side=tk.LEFT, padx=10)
    progress = ttk.Progressbar(btn_row, mode="indeterminate", length=220)

    run_btn = ttk.Button(btn_row, text="Run")
    run_btn.pack(side=tk.LEFT)
    def clear_output():
        out_text.delete("1.0", tk.END)
    ttk.Button(btn_row, text="Clear Output", command=clear_output).pack(side=tk.LEFT, padx=5)

    def safe_log(msg: str, tag: str = None, is_markdown: bool = False):
        def append():
            applied_tag = tag
            low = msg.lower()
            if applied_tag is None:
                if "error" in low or "❌" in msg or "failed" in low:
                    applied_tag = "error"
                elif low.startswith("starting") or "processing batch" in low:
                    applied_tag = "info"
                elif "suggested fix" in low or "gemini debugging output" in low:
                    applied_tag = "section"
                elif "saved" in low or "✅" in msg or "done" in low:
                    applied_tag = "success"

            if is_markdown:
                try:
                    render_markdown(msg)
                except Exception:
                    out_text.insert(tk.END, msg + ("\n" if not msg.endswith("\n") else ""), (applied_tag,) if applied_tag else ())
            else:
                msg_to_insert = msg if msg.endswith("\n") else msg + "\n"
                if applied_tag:
                    out_text.insert(tk.END, msg_to_insert, (applied_tag,))
                else:
                    out_text.insert(tk.END, msg_to_insert)

            out_text.see(tk.END)
            out_text.update_idletasks()
        app.after(0, append)

    def ui_yes_no(question: str) -> bool:
        result_holder = {"ans": False}
        gate = threading.Event()
        def ask():
            try:
                ans = messagebox.askyesno("Apply Fix?", question, parent=app)
            except Exception:
                ans = False
            result_holder["ans"] = bool(ans)
            gate.set()
        app.after(0, ask)
        gate.wait()
        return result_holder["ans"]

    def pulse_header(step: int = 0):
        if not is_running.get():
            header.configure(bg="#2b2d42")
            title_lbl.configure(bg="#2b2d42")
            status_lbl.configure(bg="#2b2d42")
            return
        colors = ["#2b2d42", "#31344f"]
        c = colors[step % len(colors)]
        header.configure(bg=c)
        title_lbl.configure(bg=c)
        status_lbl.configure(bg=c)
        app.after(300, lambda: pulse_header(step + 1))

    def set_running(running: bool):
        is_running.set(running)
        state = "disabled" if running else "normal"
        for w in (path_entry, browse_btn, json_chk, autofix_chk, apply_all_chk, prompt_entry, max_entry, run_btn):
            try:
                w.configure(state=state)
            except Exception:
                pass
        if running:
            status_lbl.configure(text="Running...", fg="#ffd166")
            run_btn.configure(text="Running...")
            progress_placeholder.pack_forget()
            progress.pack(side=tk.LEFT, padx=10)
            progress.start(12)
            pulse_header(0)
        else:
            status_lbl.configure(text="Idle", fg="#a8dadc")
            run_btn.configure(text="Run")
            try:
                progress.stop()
            except Exception:
                pass
            progress.pack_forget()
            progress_placeholder.pack(side=tk.LEFT, padx=10)

    def on_run():
        folder = path_var.get().strip()
        export_json = bool(json_var.get())
        autofix = bool(autofix_var.get())
        apply_all = bool(apply_all_var.get())
        userprompt = userprompt_var.get().strip() or None
        try:
            max_chars = int(maxchars_var.get())
        except Exception:
            max_chars = 200_000
            maxchars_var.set(max_chars)

        if not folder:
            safe_log("Please select a project folder.\n")
            return

        set_running(True)

        def worker():
            try:
                interactive = bool(autofix and not apply_all)
                prompt_func = ui_yes_no if interactive else None
                run_pipeline(folder, export_json, autofix, apply_all, userprompt, max_chars, ui_logger=safe_log, interactive=interactive, prompt_func=prompt_func)
            except Exception as e:
                safe_log(f"Error: {e}\n")
            finally:
                app.after(0, lambda: set_running(False))

        threading.Thread(target=worker, daemon=True).start()

    run_btn.configure(command=on_run)
    app.mainloop()


def main():
    if len(sys.argv) == 1:
        launch_ui()
        return

    clear_screen()
    print_logo()

    parser = argparse.ArgumentParser(description='AI Multi-File Code Debugger + Auto Fixer + User Prompt')
    parser.add_argument('path', help='Path to project directory')
    parser.add_argument('--json', action='store_true', help='Export review as JSON')
    parser.add_argument('--autofix', action='store_true', help='Enable interactive auto-fix')
    parser.add_argument('--apply-all', action='store_true', help='Apply all fixes automatically')
    parser.add_argument('--userprompt', '-p', type=str, help='Custom user requirement prompt')
    parser.add_argument('--max-chars', type=int, default=200_000, help='Max characters per Gemini batch')
    args = parser.parse_args()

    run_pipeline(args.path, args.json, args.autofix, args.apply_all, args.userprompt, args.max_chars, ui_logger=None)


if __name__ == "__main__":
    main()